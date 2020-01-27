import os
from io import BytesIO

from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLACK


def gen_batch_xlsx(delivery, data):
    all_delivery = {
        'SF': gen4sf,
        'ZT': gen4zt,
    }

    if delivery not in all_delivery:
        raise AttributeError('delivery not supported')

    return all_delivery[delivery](data)


def gen_full_file_path(file):
    pre_fix = 'conf/batch_templates/'
    return os.path.join(settings.BASE_DIR, pre_fix, file)


def central_black(sheet, value):
    font = Font(color=BLACK)
    alignment = Alignment(horizontal='center')
    sheet.value = value
    sheet.font = font
    sheet.alignment = alignment


def gen4sf(data):
    f = {
        "receiver_name": "B",
        "company_name": "C",
        "receiver_phone": "D",
        "receiver_full_address": "F",
    }
    offset = 6
    file_path = gen_full_file_path('SF.xlsx')

    virtual_workbook = BytesIO()
    wb = openpyxl.load_workbook(file_path)
    sh = wb[wb.sheetnames[0]]

    for index, value in enumerate(data):
        idx = index + offset

        central_black(sh[f'{f["receiver_name"]}{idx}'], value['receiver_name'])
        central_black(sh[f'{f["company_name"]}{idx}'], value['order_sn'])
        central_black(sh[f'{f["receiver_phone"]}{idx}'], value['receiver_phone'])
        central_black(sh[f'{f["receiver_full_address"]}{idx}'], value['receiver_full_address'])

    wb.save(virtual_workbook)
    output = virtual_workbook.getvalue()

    return output


def gen4zt(data):
    f = {
        "receiver_name": "A",
        "receiver_phone": "B",
        "receiver_full_address": "D",
        "goods_name": "E",
    }
    offset = 2
    file_path = gen_full_file_path('ZT.xlsx')

    virtual_workbook = BytesIO()
    wb = openpyxl.load_workbook(file_path)
    sh = wb[wb.sheetnames[0]]

    for index, value in enumerate(data):
        idx = index + offset

        central_black(sh[f'{f["receiver_name"]}{idx}'], value['receiver_name'])
        central_black(sh[f'{f["receiver_phone"]}{idx}'], value['receiver_phone'])
        central_black(sh[f'{f["receiver_full_address"]}{idx}'], value['receiver_full_address'])
        central_black(sh[f'{f["goods_name"]}{idx}'], f"{value['goods_name']} ({value['order_sn']})")

    wb.save(virtual_workbook)
    output = virtual_workbook.getvalue()

    return output
